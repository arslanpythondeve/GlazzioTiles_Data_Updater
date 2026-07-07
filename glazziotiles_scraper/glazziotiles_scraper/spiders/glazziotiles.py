import json, os
import csv, glob
from datetime import datetime
from typing import Iterable, Any
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl import load_workbook
from scrapy import Spider, Request


class GlazzioTilesSpider(Spider):
    name = "glazziotiles"
    start_url = "https://www.glazziotiles.com/"

    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
        'priority': 'u=0, i',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36',
    }

    custom_settings = {
        "CONCURRENT_REQUESTS": 4,
        # 'DOWNLOAD_DELAY': 2,
        # 'RANDOMIZE_DOWNLOAD_DELAY': True,

        'RETRY_TIMES': 5,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 408, 429, 410],

        "ITEM_PIPELINES": {
                "glazziotiles_scraper.pipelines.GlazzioTilesImagesPipeline": 1,
            },

        "IMAGES_STORE": "output/images",
    }

    def __init__(self, **kwargs: Any):
        super().__init__(**kwargs)
        self.input_products = self.read_input_file()
        self.config = self.read_json_file('input/config.json')
        self.input_sku_column_name = self.config.get('excel_sku_column_name', '')
        self.input_headers = list(self.input_products[0].keys()) if self.input_products else []
        self.input_skus = [p.get(self.input_sku_column_name, "") for p in self.input_products]
        self.spec_fields = set()
        self.scraped_skus = set()
        self.items = []


    def start_requests(self) -> Iterable[Any]:
        yield Request(url=self.start_url, headers=self.headers, callback=self.parse)

    def parse(self, response, **kwargs):
        for sku in self.input_skus:

            if sku in self.scraped_skus:
                self.logger.info(f"Skipping duplicate SKU: {sku}")
                continue

            product_url = f"https://www.glazziotiles.com/default.aspx?page=item%20detail&itemcode={sku}"

            yield Request(url=product_url, headers=self.headers, callback=self.parse_details, meta={'sku': sku})

    def parse_details(self, response):
        sku_number = response.meta.get('sku', '')

        item = OrderedDict()

        product = next((p for p in self.input_products if p.get(self.input_sku_column_name, '') == sku_number), {})

        for col in self.input_headers:
            item[f'_{col}'] = product.get(col, "")

        item['-'] = ''
        item['--'] = ''

        for specs in response.css('td.ItemDetailTopAlign tr'):
            specs_key = specs.css('.ItemDetailattribute_hdr::text').get('').strip()
            specs_value = specs.css('.ItemDetailattribute::text').get('').strip() or specs.css('#txtItemDetailQuantity0::attr(value)').get('').strip()

            item[specs_key] = specs_value


        carousel_images = [f"https://www.glazziotiles.com{img}" for img in
                           response.css("#mycarousel li img::attr(src)").getall() if img and img.strip()]

        if carousel_images:
            images = carousel_images
        else:
            img = response.css("#imgItemDetail::attr(src)").get('')
            images = [f"https://www.glazziotiles.com{img}"] if img and img.strip() else []

        item["image_urls"] = images

        try:
            secondary_specs = response.xpath('//td[@class="ItemDetailImageTD"]/parent::tr/following-sibling::tr[2]/td//text()').getall()

            for spec in secondary_specs:
                spec = spec.strip()

                if not spec:
                    continue

                if ':' in spec:
                    key, value = spec.split(':', 1)
                    item[key.strip()] = value.strip()

        except Exception as e:
            self.logger.warning(f"Failed to parse secondary specs: {e}")

        item['Url'] = response.url

        self.scraped_skus.add(sku_number)
        self.items.append(item)
        yield item

    def read_input_file(self, sheet_name=None):
        input_files = sorted(glob.glob("input/*.csv") + glob.glob("input/*.xls") + glob.glob("input/*.xlsx"))
        if not input_files:
            self.logger.error("\n\nNo Excel/CSV file exists inside the input folder\n")
            exit(0)

        file_path = input_files[0]

        try:
            ext = os.path.splitext(file_path)[1].lower()

            if ext == ".csv":
                with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.DictReader(f)
                    return list(reader)

            elif ext in [".xlsx", ".xls"]:
                wb = load_workbook(file_path, data_only=True)
                ws = wb[sheet_name] if sheet_name else wb.active

                rows = list(ws.iter_rows(values_only=True))

                if not rows:
                    return []

                headers = rows[0]

                data = []
                for row in rows[1:]:
                    data.append(dict(zip(headers, row)))

                return data

            else:
                self.logger.error(f"Unsupported file format: {ext}")
                return []

        except Exception as e:
            self.logger.error(f"Error reading input file ({file_path}): {e}")
            return []

    def read_json_file(self, file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                return json.load(f)

        except Exception as e:
            self.logger.error(f"Error reading JSON file: {e}")

        return {}

    def closed(self, reason):
        missing_skus = set(self.input_skus) - self.scraped_skus

        for sku in missing_skus:
            product = next((p for p in self.input_products if p.get(self.input_sku_column_name, "") == sku), {})

            item = OrderedDict()

            for col in self.input_headers:
                item[f'_{col}'] = product.get(col, "")

            self.items.append(item)

        if not self.items:
            self.logger.info("No items scraped.")
            return

        headers = []

        for item in self.items:
            for key in item.keys():
                if key not in headers:
                    headers.append(key)

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(headers)

        for item in self.items:
            row = []
            for col in headers:
                value = item.get(col, "")

                if isinstance(value, list):
                    value = ", ".join(value)

                row.append(value)

            ws.append(row)

        output_file = f'output/GlazzioTiles_{datetime.now().strftime("%d%m%Y%H%M")}.xlsx'
        wb.save(output_file)

        self.logger.info(f"Saved {len(self.items)} items to {output_file}")
